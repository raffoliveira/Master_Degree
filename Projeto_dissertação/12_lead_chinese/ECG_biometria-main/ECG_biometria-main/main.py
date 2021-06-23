'''
KEY:
1 - Utiliza um canal por amostra e gera amostras com as 12 derivações.
2 - Utiliza as 12 derivações em uma única amostra. Detecta o pico no LEAD I, mas a amostra é do tipo (12,400)
3 - Utiliza um canal por amostra e gera amostras somente com a derivação II.

'''
KEY = 1

if KEY == 1:
    from utils_arrhythmiasDB_1 import create_files, load_training_data, load_test_data
if KEY == 2:
    from utils_arrhythmiasDB_2 import create_files, load_training_data, load_test_data
if KEY == 3:
    from utils_arrhythmiasDB_3 import create_files, load_training_data, load_test_data


from keras import callbacks
from keras.models import Model
import matplotlib.pyplot as plt
import time
import os
from openpyxl import load_workbook

from model import cnn_model
from metrics import *


def scheduler(epoch):
    lrs = [0.01] * 40 + [0.001] * 20 + [0.0001] * 20
    return lrs[epoch]


if __name__ == '__main__':
    sizeBeat = 800
    filter = 0
    useDataAugmentation = False
    singleFlag = False
    epochs = 40
    path = '../../results_arrDB3/'

    if not os.path.isdir(path):
        os.mkdir(path)

    #                       GET A DICTIONARY WITH THE ACRONYMS OF THE RHYTHMS
    ##########################################################################################################
    rhythm_map = {}
    workbook = load_workbook(filename=r'../../dataset2/RhythmNames.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2,
                               min_col=0,
                               values_only=True):
        key = row[0].strip(' ')
        rhythm_map[key] = row[1]
    ##########################################################################################################


    if os.path.isfile(path + 'trainData.npy'):
        train = np.load(path + 'trainData.npy', allow_pickle=True)
        test = np.load(path + 'testData.npy', allow_pickle=True)
    else:
        train, test = create_files(sizeBeat=800, filter=0, useDataAugmentation=useDataAugmentation, path=path)

    print('\n\tvalid SR segments = ' + str(len(train)))
    x_train, y_train, labels = load_training_data(train_dict=train, useDataAugmentation=useDataAugmentation)

    input_dim = (x_train.shape[1], x_train.shape[2])
    classes = len(np.unique(labels))
    model = cnn_model(input_dim, num_classes=classes)

    LRscheduler = callbacks.LearningRateScheduler(scheduler, verbose=1)

    print('\n\tStarting CNN training:\n')
    t1 = time.time()
    history = model.fit(x_train, y_train, batch_size=100, epochs=epochs, callbacks=[LRscheduler], validation_split=0.1)

    t2 = time.time()
    time_seconds = t2 - t1
    print('\n\tTime spent for training: {}s'.format(time_seconds))

    plt.plot(history.history['accuracy'])
    plt.plot(history.history['val_accuracy'])
    plt.title('Model accuracy')
    plt.ylabel('accuracy')
    plt.xlabel('epoch')
    plt.legend(['train', 'validation'], loc='upper left')
    plt.savefig(path + r'ACC.png', format='png')
    plt.show()
    plt.clf()

    # summarize history for loss
    plt.plot(history.history['loss'])
    plt.plot(history.history['val_loss'])
    plt.title('Model loss')
    plt.ylabel('loss')
    plt.xlabel('epoch')
    plt.legend(['train', 'validation'], loc='upper left')
    plt.savefig(path + r'LOSS.png', format='png')
    plt.show()
    plt.clf()

    model.save('{}ECG-model'.format(path))

    ver_model = Model(inputs=model.input,
                      outputs=model.get_layer('FC2').output)

    del train, x_train, y_train
    for rhythm in rhythm_map.keys():
        try:
            x_test, y_test, labels = load_test_data(test, rhythm, rhythm_map)
            print('\n\n\tTest with: {} ({}) [samples={}]'.format(rhythm_map[rhythm], rhythm,  len(labels)))
            features = ver_model.predict(x_test)
            genuine_scores, impostor_scores = get_scores(features, features, labels, labels)
            d = calc_decidability(genuine_scores, impostor_scores)
            eer, threshold, fm, fnm = calc_eer(genuine_scores, impostor_scores,  path=path, plot_det=False)

            print('\tEER = ' + str(eer))
            print('\tDecidability = ' + str(d))
            print('\tGenuine pairs = ' + str(len(genuine_scores)))
            print('\tImpostors pairs = ' + str(len(impostor_scores)))
            print('\tFalse match samples = ' + str(fm))
            print('\tFalse non match samples = ' + str(fnm))
        except:
            print('>\tIndividuals with heart rhythm ', rhythm, ' did not generate valid samples')
