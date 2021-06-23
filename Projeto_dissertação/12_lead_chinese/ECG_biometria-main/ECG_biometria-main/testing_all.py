from keras.models import load_model
from keras.models import Model
import os
from openpyxl import load_workbook

from metrics import *
from utils_arrhythmiasDB_3 import create_files, load_training_data, load_test_data


if __name__ == '__main__':
    sizeBeat = 800
    filter = 0
    path = '../../results_arrDB3/'

    if not os.path.isdir(path):
        os.mkdir(path)

    #                       GET A DICTIONARY WITH THE ACRONYMS OF THE RHYTHMS
    ##########################################################################################################
    rhythm_map = {}
    workbook = load_workbook(filename=r'../../dataset2/RhythmNames.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2,
                               min_col=0,
                               values_only=True):
        key = row[0].strip(' ')
        rhythm_map[key] = row[1]
    ##########################################################################################################

    test = np.load(path + 'testData.npy', allow_pickle=True)
    model = load_model(path + 'ECG-model')
    ver_model = Model(inputs=model.input,
                      outputs=model.get_layer('FC2').output)

    del model
    all_features = []
    all_labels = []
#    del rhythm_map['SB'], rhythm_map['AFIB'], rhythm_map['ST']
    print('\n')
    for rhythm in rhythm_map.keys():
        try:
           x_test, y_test, labels = load_test_data(test, rhythm, rhythm_map)
           print('\tPrcessing ' + rhythm)
           for i in range(len(x_test)):
                feature = ver_model.predict(x_test[i].T)
                all_features.append(feature)
                all_labels.append(labels[i])
        except:
            print('\n\t>Individuals with heart rhythm ', rhythm, ' did not generate valid samples')
    del test, x_test, y_test, labels

    all_features = np.asarray(all_features)
    all_features = np.squeeze(all_features)
#    print(all_features.shape)
    print('\n\n\tEvaluating model with all valid samples:')

    genuine_scores, impostor_scores = get_scores(all_features, all_features, all_labels, all_labels)
    d = calc_decidability(genuine_scores, impostor_scores)
    eer, threshold, fm, fnm = calc_eer(genuine_scores, impostor_scores,  path='./test/', plot_det=False)

    print('\tEER = ' + str(eer))
    print('\tDecidability = ' + str(d))
    print('\tGenuine pairs = ' + str(len(genuine_scores)))
    print('\tImpostors pairs = ' + str(len(impostor_scores)))
    print('\tFalse match samples = ' + str(fm))
    print('\tFalse non match samples = ' + str(fnm))
