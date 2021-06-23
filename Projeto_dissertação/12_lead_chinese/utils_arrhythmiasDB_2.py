#2 - Utiliza as 12 derivações em uma única amostra. Detecta o pico no LEAD I, mas a amostra é do tipo (12,400)

import numpy as np
from scipy.signal import firwin, firls, iirdesign, filtfilt, butter
from sklearn.metrics.pairwise import euclidean_distances
import os
from openpyxl import load_workbook
import glob
from keras.utils.np_utils import to_categorical
import matplotlib.pyplot as plt
from ecgdetectors import Detectors
from detectors import *
import csv
import pickle

def read_csv(file):
    with open(file, 'r') as f:
        reader = csv.reader(f)
        data = [row for row in reader]
        data = np.delete(np.array(data), 0, axis=0)
        data = data.astype('float32')
    return data


def filter_data(data, filter, freq):
    bpFilt = firwin(5, [0.5, 40], pass_zero=False, fs=freq) #perguntar pedro se sigo o do codigo ou do artigo
    dataFilt = filtfilt(bpFilt, 1, data)

    #Signal between 0 and 1
    dataFilt = dataFilt - np.mean(dataFilt, axis=1, keepdims=True)
    dataFilt = dataFilt + np.abs(np.amin(dataFilt, axis=1, keepdims=True))
    dataFilt = dataFilt / np.std(dataFilt, axis=1, keepdims=True)
    dataFilt = dataFilt / np.amax(dataFilt, axis=1, keepdims=True)

    #detectors = Detectors(freq)
    #qrs_i = [detectors.engzee_detector(dataFilt[i,:]) for i in range(12)]
    qrs_i = [detect(signal=dataFilt[i,:], rate=500) for i in range(12)]

    '''
    print(qrs_i[0])
    print(dataFilt[0][3])
    y = [dataFilt[0][i] for i in qrs_i[0]]
    plt.plot(np.arange(5000), dataFilt[0])
    plt.plot(qrs_i[0], y, 'o')
    plt.show()
    '''

    if filter != 0:
        if filter == 1:
            data = dataFilt
        else:
            a = 1
            if filter == 2:
                b = firwin(4, [1, 99], pass_zero=False, fs=freq)
            elif filter ==3:
                b = firwin(4, [0.5, 99], pass_zero=False, fs=freq)
            elif filter == 4:
                b = firwin(20, [0.5, 50], pass_zero=False, fs=freq)
            elif filter == 5:
                b = firls(21, [0.5, 1, 40, 45], desired=[0,1,1,0], weight=[1,2], fs=freq)
            elif filter == 6:
                b, a = iirdesign(wp=[1,50], ws=[0.5,55], gstop=45, gpass=1, ftype='ellip', fs=freq)
            elif filter == 7:
                b = firls(21, [0.5, 1, 30, 35], desired=[0, 1, 1, 0], weight=[1, 2], fs=freq)
            elif filter == 8:
                b = firls(81, [0.5, 1, 30, 35], desired=[0,1,1,0], weight=[1,2], fs=freq)
            elif filter == 9:
                b, a = iirdesign(wp=[0.5, 80], ws=[0.3, 95], gstop=45, gpass=1, ftype='ellip', fs=freq)
            elif filter == 10:
                b, a = iirdesign(wp=[4, 20], ws=[3, 25], gstop=45, gpass=1, ftype='ellip', fs=freq)

            data = filtfilt(b, a, data)

    return data, dataFilt, qrs_i


def read_raw_arrhythmiasDB(filter, path):
    freq = 500
    dir = r'../../dataset2/ECGData/'
    files = glob.glob(dir + '/*.csv')
    files = [file.replace('\\', '/') for file in files]
    qrs_dict = {}

    if not os.path.isdir(path + 'data'):
        os.mkdir(path + 'data')
    if not os.path.isdir(path + 'dataFilt'):
        os.mkdir(path + 'dataFilt')

    for id, file_path in enumerate(files):
       # if id<300:
            print('\tProcessing individual {}'.format(id))
            *rest, file = file_path.split('/')
            file, ext = file.split('.')
            data = read_csv(file_path)
            data = data.T
            data, dataFilt, qrs_i = filter_data(data, filter, freq)

            # Signal between 0 and 1
            data = data - np.mean(data, axis=1, keepdims=True)
            data = data + np.abs(np.amin(data, axis=1, keepdims=True))
            data = data / np.std(data, axis=1, keepdims=True)
            data = data / np.amax(data, axis=1, keepdims=True)

            qrs_dict[file] = qrs_i

            np.save('{}data/{}.npy'.format(path, file), data)
            np.save('{}dataFilt/{}.npy'.format(path, file), dataFilt)

    outfile = open(path + 'qrs_dict', 'wb')
    pickle.dump(qrs_dict, outfile)
    outfile.close()


def outlier_removal(data_paths, dataFilt_paths, qrs_dict, ids_rhythm, rhythm_map,  sizeBeat, path):
    freqSample = 500
    sizeBeat = int(sizeBeat/2)
    ids = len(data_paths)
    ann_valid_dict = {}
    ii = 0 #SELECIONA QUAL DERIVAÇÃO USAR DE REFERÊNCIA

    if not os.path.isdir(path + 'train_ecg'):
        os.mkdir(path + 'train_ecg')
    if not os.path.isdir(path + 'train_ecg_t'):
        os.mkdir(path + 'train_ecg_t')

    for id in range(ids):
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')
        rhythm = ids_rhythm[id_file]

        ecgDII_t = np.load(path + 'data/' + id_file + '.npy')
        ecgDII = np.load(path + 'dataFilt/' + id_file + '.npy')

        anns = qrs_dict[id_file]
        ann_valid = np.zeros(len(anns[ii]))
        train_ecg = []
        train_ecg_t = []


        for jj in range(len(anns[ii])):
            bwnd = freqSample/2 - 1    #200ms   #DUVIDA
            fwnd = sizeBeat - bwnd

            if anns[ii][jj] - bwnd < 0:
                startPoint = 0
            else:
                startPoint = anns[ii][jj] - bwnd

            if anns[ii][jj] + fwnd >= ecgDII.shape[1]:
                endPoint = ecgDII.shape[1]
            else:
                endPoint = anns[ii][jj] + fwnd

            startPoint = int(startPoint)
            endPoint = int(endPoint)

            beatWaveDII = ecgDII[:, startPoint:endPoint]
            beatWaveDII_t = ecgDII_t[:, startPoint:endPoint]

            if beatWaveDII.shape[1] == sizeBeat:
                valid = 1
                #criterio 1: se valor maximo nao for o QRS, exclui
                max_pos = np.argmax(beatWaveDII[0, :])

                if max_pos <= (bwnd-3):
                    valid = 0
                if max_pos >= (bwnd+3): ### duvida
                    valid = 0

                ann_valid[jj] = valid

                if valid == 1:
                    train_ecg.append(beatWaveDII)
                    train_ecg_t.append(beatWaveDII_t)

        train_ecg = np.asarray(train_ecg)
        train_ecg_t = np.asarray(train_ecg_t)

        ann_valid_dict[id_file] = ann_valid

        np.save(path + 'train_ecg/' + id_file +'.npy', train_ecg)
        np.save(path + 'train_ecg_t/' + id_file + '.npy', train_ecg_t)

    with open(path + 'ann_valid', 'wb') as outfile:
        pickle.dump(ann_valid_dict, outfile)


    print('\n\tScale the beatles according to Sinus Rhythm (normal rhythm) ...')

    all_m1 = []
    all_m1_t = []

    for id in range(ids):
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')
        rhythm = ids_rhythm[id_file]

        train_ecg_t = np.load(path + 'train_ecg_t/' + id_file + '.npy', allow_pickle=True)

        samples = np.sum([len(i) for i in train_ecg_t])

        if rhythm == 'SR' and samples>0:
            data = np.load(path + 'data/' + id_file + '.npy', allow_pickle=True)
            dataFilt = np.load(path + 'dataFilt/' + id_file + '.npy', allow_pickle=True)
            train_ecg = np.load(path + 'train_ecg/' + id_file + '.npy', allow_pickle=True)


            m1_t = np.mean(train_ecg_t, axis=-1)
            m1_t = m1_t[np.isfinite(m1_t)]
            m1_mean_t = np.mean(m1_t)
            data = data - m1_mean_t

            m1 = np.mean(train_ecg, axis=-1)
            m1 = m1[np.isfinite(m1)]
            m1_mean = np.mean(m1)
            dataFilt = dataFilt - m1_mean

            all_m1_t.append(m1_t)
            all_m1.append(m1)

            np.save(path + 'data/' + id_file + '.npy', data)
            np.save(path + 'dataFilt/' + id_file + '.npy', dataFilt)

    print('\t> averages calculated')

    for id in range(ids):
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')
        rhythm = ids_rhythm[id_file]

        data = np.load(path + 'data/' + id_file + '.npy', allow_pickle=True)
        dataFilt = np.load(path + 'dataFilt/' + id_file + '.npy', allow_pickle=True)
        train_ecg = np.load(path + 'train_ecg/' + id_file + '.npy', allow_pickle=True)
        train_ecg_t = np.load(path + 'train_ecg_t/' + id_file + '.npy', allow_pickle=True)

        samples = np.sum([len(i) for i in train_ecg_t])

        if rhythm != 'SR' and samples > 0:
            m2_t = np.mean(train_ecg_t, axis=-1)
            m2_t = m2_t[np.isfinite(m2_t)]
            m2_mean_t = np.mean(m2_t)

            m1_max_t = np.amax([np.amax(i) for i in all_m1_t])
            m1_min_t = np.amin([np.amin(i) for i in all_m1_t])
            dif1_t = m1_max_t - m1_min_t

            m2_max_t = np.amax(m2_t)
            m2_min_t = np.amin(m2_t)
            dif2_t = m2_max_t - m2_min_t

            data = data - m2_mean_t
            #data = ((dif1_t/dif2_t)*data)

            m2 = np.mean(train_ecg, axis=-1)
            m2 = m2[np.isfinite(m2)]
            m2_mean = np.mean(m2)

            m1_max = np.amax([np.amax(i) for i in all_m1])
            m1_min = np.amin([np.amax(i) for i in all_m1])
            dif1 = m1_max - m1_min

            m2_max = np.amax(m2)
            m2_min = np.amin(m2)
            dif2 = m2_max - m2_min

            dataFilt = dataFilt - m2_mean
            #dataFilt = ((dif1 / dif2) * dataFilt)

            np.save(path + 'data/' + id_file + '.npy', data)
            np.save(path + 'dataFilt/' + id_file + '.npy', dataFilt)

    print('\t> signals are normalized\n')


    for id in range(ids):
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')
        rhythm = ids_rhythm[id_file]

        ecgDII = np.load(path + 'dataFilt/' + id_file + '.npy')

        anns = qrs_dict[id_file]
        ann_valid = np.zeros(len(anns[ii]))
        train_ecg = []

        for jj in range(len(anns[ii])):
            bwnd = freqSample/2 - 1    #200ms   #DUVIDA
            fwnd = sizeBeat - bwnd

            if anns[ii][jj] - bwnd < 0:
                startPoint = 0
            else:
                startPoint = anns[ii][jj] - bwnd

            if anns[ii][jj] + fwnd >= ecgDII.shape[1]:
                endPoint = ecgDII.shape[1]
            else:
                endPoint = anns[ii][jj] + fwnd

            startPoint = int(startPoint)
            endPoint = int(endPoint)

            beatWaveDII = ecgDII[:, startPoint:endPoint]

            if beatWaveDII.shape[1] == sizeBeat:
                valid = 1
                #criterio 1: se valor maximo nao for o QRS, exclui
                max_pos = np.argmax(beatWaveDII[0, :])

                if max_pos <= (bwnd-3):
                    valid = 0
                if max_pos >= (bwnd+3): ### duvida
                    valid = 0

                ann_valid[jj] = valid

                if valid == 1:
                    train_ecg.append(beatWaveDII)

        train_ecg = np.asarray(train_ecg)
        ann_valid_dict[id_file] = ann_valid

        np.save(path + 'train_ecg/' + id_file +'.npy', train_ecg)

    with open(path + 'ann_valid', 'wb') as outfile:
        pickle.dump(ann_valid_dict, outfile)


    print('\tCalculate distances...\n')
    d_mean = {}
    for id in range(ids):
        d_mean[id] = []
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')

        rhythm = ids_rhythm[id_file]
        anns = qrs_dict[id_file]

        train_ecg = np.load(path + 'train_ecg/' + id_file + '.npy', allow_pickle=True)
        ecgDII = np.load(path + 'dataFilt/' + id_file + '.npy', allow_pickle=True)

        mean_ecg = np.mean(train_ecg, axis=0)

        del train_ecg

        ann_valid = np.zeros(len(anns[ii]))

        for jj in range(len(anns[ii])):
            bwnd = freqSample / 2 - 1
            fwnd = sizeBeat - bwnd

            if anns[ii][jj] - bwnd < 0:
                startPoint = 0
            else:
                startPoint = anns[ii][jj] - bwnd

            if anns[ii][jj] + fwnd >= ecgDII.shape[1]:
                endPoint = ecgDII.shape[1]
            else:
                endPoint = anns[ii][jj] + fwnd

            startPoint = int(startPoint)
            endPoint = int(endPoint)

            beatWaveDII = ecgDII[:, startPoint:endPoint]

            if beatWaveDII.shape[1] == sizeBeat:
                valid = 1
                # criterio 1: se valor maximo nao for o QRS, exclui
                max_pos = np.argmax(beatWaveDII[0, :])

                if max_pos <= (bwnd - 3):
                    valid = 0
                if max_pos >= (bwnd + 3):
                    valid = 0

                ann_valid[jj] = valid

                if valid == 1:
                    try:
                        dist = euclidean_distances(beatWaveDII, mean_ecg)
                        d_mean[id].append(dist)
                    except:
                        ann_valid[jj] = 0

        ann_valid_dict[id_file] = ann_valid

    with open(path + 'ann_valid', 'wb') as outfile:
        pickle.dump(ann_valid_dict, outfile)


    print('\tRemoving outliers...\n')
    for id in range(ids):
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')

        rhythm = ids_rhythm[id_file]
        anns = qrs_dict[id_file]

        train_ecg = np.load(path + 'train_ecg/' + id_file + '.npy', allow_pickle=True)
        ecgDII = np.load(path + 'dataFilt/' + id_file + '.npy', allow_pickle=True)

        mean_ecg = np.mean(train_ecg, axis=0)
        ann_valid = np.zeros(len(anns[ii]))

        for jj in range(len(anns[ii])):
            bwnd = freqSample / 2 - 1
            fwnd = sizeBeat - bwnd

            if anns[ii][jj] - bwnd < 0:
                startPoint = 0
            else:
                startPoint = anns[ii][jj] - bwnd

            if anns[ii][jj] + fwnd >= ecgDII.shape[1]:
                endPoint = ecgDII.shape[1]
            else:
                endPoint = anns[ii][jj] + fwnd

            startPoint = int(startPoint)
            endPoint = int(endPoint)

            beatWaveDII = ecgDII[:, startPoint:endPoint]

            if beatWaveDII.shape[1] == sizeBeat:
                # criterio 1: se valor maximo nao for o QRS, exclui
                max_pos = np.argmax(beatWaveDII[0, :])
                ann_valid[jj] = 1

                if max_pos <= (bwnd - 3):
                    ann_valid[jj] = 0
                if max_pos >= (bwnd + 3):
                    ann_valid[jj] = 0

                if ann_valid[jj] == 1:
                    try:
                        d = np.mean(np.abs(euclidean_distances(beatWaveDII, mean_ecg)))
                        if d > np.abs(np.mean(d_mean[id]) + 2.5 * np.std(d_mean[id])):
                            ann_valid[jj] = 0
                    except:
                        ann_valid[jj] = 0

        ann_valid_dict[id_file] = ann_valid

    with open(path + 'ann_valid', 'wb') as outfile:
        pickle.dump(ann_valid_dict, outfile)

    print('\t> done')

    t = {}
    v = {}
    for key in rhythm_map.keys():
        t[key] = 0
        v[key] = 0

    #with open(path + 'ann_valid', 'rb') as infile:
    #    ann_valid = pickle.load(infile)

    for id in range(ids):
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')

        rhythm = ids_rhythm[id_file]

        ann = qrs_dict[id_file]
        total = len(ann[ii])

        ann_valid = ann_valid_dict[id_file]
        valids = np.sum(ann_valid)

        t[rhythm] = t[rhythm] + total
        v[rhythm] = v[rhythm] + valids

    for key in rhythm_map.keys():
        print('\n\tRhythm: {} [{}]'.format(rhythm_map[key], key))
        print('\tTotal: {} \tValids: {}'.format(t[key], v[key]))


def createIMDBStructure(data_paths, qrs_dict, ids_rhythm, sizeBeat, useDataAugmentation, path):
    sizeBeat = int(sizeBeat/2)
    freqSample = 500
    thbc = 1    # total heart beat count
    phbc = 1
    ids = len(data_paths)
    bwnd = freqSample/2 - 1
    fwnd = sizeBeat - bwnd

    ii = 0
    dataset = []
    classes = 0

    with open(path + 'ann_valid', 'rb') as infile:
        ann_valid_dict = pickle.load(infile)

    label_id = -1
    for id in range(ids):
        aux = 0
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')

        rhythm = ids_rhythm[id_file]

        if rhythm == 'SR':
            ecgDII = np.load(path + 'data/' + id_file + '.npy')
            anns = qrs_dict[id_file]
            ann_valid = ann_valid_dict[id_file]

            numHB = np.sum(ann_valid)

            if numHB > 0:
                classes = classes + 1

            for jj in range(len(anns[ii])):
                sample = {}

                if anns[ii][jj] - bwnd < 0:
                    startPoint = 0
                else:
                    startPoint = anns[ii][jj] - bwnd

                if anns[ii][jj] + fwnd >= ecgDII.shape[1]:
                    endPoint = ecgDII.shape[1]
                else:
                    endPoint = anns[ii][jj] + fwnd

                startPoint = int(startPoint)
                endPoint = int(endPoint)

                beatWaveDII = ecgDII[:, startPoint:endPoint].T

                if phbc >= numHB - 10:
                    set = 2
                else:
                    set = 1

                if beatWaveDII.shape[0] == sizeBeat and ann_valid[jj] == 1:
                    if aux == 0:
                        label_id = label_id + 1
                        aux = 1
                        print('\tId {} has valid segments for training'.format(id))
                    sample['data'] = beatWaveDII
                    sample['label'] = label_id
                    sample['thbc'] = thbc
                    sample['set'] = set
                    sample['da'] = 0
                    thbc = thbc + 1
                    dataset.append(sample)

                    if useDataAugmentation:
                        sample['data'] = 1.1 * beatWaveDII
                        sample['label'] = label_id
                        sample['thbc'] = thbc
                        sample['set'] = set
                        sample['da'] = 1
                        thbc = thbc + 1
                        dataset.append(sample)

                        sample['data'] = 0.9 * beatWaveDII
                        sample['label'] = label_id
                        sample['thbc'] = thbc
                        sample['set'] = set
                        sample['da'] = 1
                        thbc = thbc + 1
                        dataset.append(sample)

                        # atenua a onda P
                        sample['data'] = beatWaveDII * np.concatenate((np.full((60,12), 0.3), np.ones((340,12))), axis=0)
                        sample['label'] = label_id
                        sample['thbc'] = thbc
                        sample['set'] = set
                        sample['da'] = 1
                        thbc = thbc + 1
                        dataset.append(sample)

                        # da ganho a onda P
                        sample['data'] = beatWaveDII * np.concatenate((np.full((60,12), 1.2), np.ones((340,12))), axis=0)
                        sample['label'] = label_id
                        sample['thbc'] = thbc
                        sample['set'] = set
                        sample['da'] = 1
                        thbc = thbc + 1
                        dataset.append(sample)

                        # atenua a onda T
                        sample['data'] = beatWaveDII * np.concatenate((np.ones((150,12)), np.full((250,12), 0.3)), axis=0)
                        sample['label'] = label_id
                        sample['thbc'] = thbc
                        sample['set'] = set
                        sample['da'] = 1
                        thbc = thbc + 1
                        dataset.append(sample)

                        # da ganho a onda T
                        sample['data'] = beatWaveDII * np.concatenate((np.ones((150,12)), np.full((250,12), 1.2)), axis=0)
                        sample['label'] = label_id
                        sample['thbc'] = thbc
                        sample['set'] = set
                        sample['da'] = 1
                        thbc = thbc + 1
                        dataset.append(sample)

                    phbc = phbc + 1

    np.save(path + 'trainData.npy', dataset)
#    print('\tTrain classes: ', classes)
    return dataset


def createTestStructure(data_paths, qrs_dict, ids_rhythm, sizeBeat, path):
    sizeBeat = int(sizeBeat / 2)
    freqSample = 500
    ids = len(data_paths)
    bwnd = freqSample / 2 - 1
    fwnd = sizeBeat - bwnd

    ii = 0
    dataset = []
    classes = dict.fromkeys(ids_rhythm.values())
    for i in classes.keys():
        classes[i] = 0

    with open(path + 'ann_valid', 'rb') as infile:
        ann_valid_dict = pickle.load(infile)

    label_id = -1
    for id in range(ids):
        aux = 0
        label_id = id
        *rest, id_file = data_paths[id].split('/')
        id_file, ext = id_file.split('.')

        rhythm = ids_rhythm[id_file]

        ann_valid = ann_valid_dict[id_file]
        numHB = np.sum(ann_valid)
        if numHB > 0:
            classes[rhythm] = classes[rhythm] + 1

        if rhythm != 'SRiiiii':
            ecgDII = np.load(path + 'data/' + id_file + '.npy')
            anns = qrs_dict[id_file]
#            ann_valid = ann_valid_dict[id_file]

#            numHB = [np.sum(i) for i in ann_valid]
#            numHB = np.sum(numHB)
#            if numHB > 0:
#                classes[rhythm] = classes[rhythm] + 1

            limiter = 0
            for jj in range(len(anns[ii])):
                sample = {}

                if anns[ii][jj] - bwnd < 0:
                    startPoint = 0
                else:
                    startPoint = anns[ii][jj] - bwnd

                if anns[ii][jj] + fwnd >= ecgDII.shape[1]:
                    endPoint = ecgDII.shape[1]
                else:
                    endPoint = anns[ii][jj] + fwnd

                startPoint = int(startPoint)
                endPoint = int(endPoint)

                beatWaveDII = ecgDII[:, startPoint:endPoint].T

                if beatWaveDII.shape[0] and ann_valid[jj] == 1 and limiter < 2000000:
                    if aux == 0:
           #             label_id = label_id + 1
                        aux = 1
                        print('\tId {} has valid segments for test with condition {}'.format(id, rhythm))
                    sample['rhythm'] = rhythm
                    sample['data'] = beatWaveDII
                    sample['label'] = label_id
                    dataset.append(sample)

                    limiter = limiter + 1

    np.save(path + 'testData.npy', dataset)
    print('\n\n\tClasses: ', classes)
    return dataset


def create_files(sizeBeat, filter, useDataAugmentation, path):
    if not os.path.isdir(path):
        os.mkdir(path)

    #         GET A DICTIONARY WITH THE ACRONYMS OF THE RHYTHMS
    rhythm_map = {}
    workbook = load_workbook(filename=r'../../dataset2/RhythmNames.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2,
                               min_col=0,
                               values_only=True):
        key = row[0].strip(' ')
        rhythm_map[key] = row[1]

    #         GET A DICTIONARY WITH THE CARDIAC RHYTHM OF EACH INDIVIDUAL
    ids_rhythm = {}
    workbook = load_workbook(filename=r'../../dataset2/Diagnostics.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2,
                               min_col=0,
                               values_only=True):
        key = row[0]
        r = row[1].strip(' ')
        if r == 'SA':
            r = 'SAAWR'
        ids_rhythm[key] = r

    #         READ RAW DATA, FILTER AND DETECT QRS
#    read_raw_arrhythmiasDB(filter=filter, path=path)

    #         START THE OUTLIER REMOVAL
    data_paths = glob.glob(path + 'data/*.npy')
    data_paths = [file.replace('\\', '/') for file in data_paths]

    dataFilt_paths = glob.glob(path + 'dataFilt/*.npy')
    dataFilt_paths = [file.replace('\\', '/') for file in dataFilt_paths]

    with open(path + 'qrs_dict', 'rb') as infile:
        qrs_dict = pickle.load(infile)

   # outlier_removal(data_paths, dataFilt_paths, qrs_dict, ids_rhythm, rhythm_map, sizeBeat, path)

    print('\n\n\tStarting heartbeat segmentation and the imdb structure creation...')
   # imdb = createIMDBStructure(data_paths, qrs_dict, ids_rhythm, sizeBeat, useDataAugmentation, path)
    print('\n\n\tStarting heartbeat segmentation and the test structure creation...')
    test = createTestStructure(data_paths, qrs_dict, ids_rhythm, sizeBeat, path)

    return imdb, test


def load_training_data(train_dict, useDataAugmentation):
    x_train = []
    y_train = []

    for sample in train_dict:
        if sample['da'] == 1:
            if useDataAugmentation:
                x_train.append(sample['data'])
                y_train.append(sample['label'])
        else:
            x_train.append(sample['data'])
            y_train.append(sample['label'])

    x_train = np.array(x_train)
    y_train = np.array(y_train).reshape(-1,1)

    classes = len(np.unique(y_train))
    return x_train, to_categorical(y_train), y_train


def load_test_data(test_dict, rhythm, rhythm_map):
    x = []
    y = []
    for sample in test_dict:
        if rhythm == sample['rhythm']:
            x.append(sample['data'])
            y.append(sample['label'])

    x = np.asarray(x)
    y = np.asarray(y).reshape(-1,1)

    return x, to_categorical(y), y


if __name__ == '__main__':
    path = '../../results_arrDB2/'
    sizeBeat = 800

    imdb, test = create_files(sizeBeat, filter=0, useDataAugmentation=False, path=path)
#    print(len(imdb))
#    print(len(test))


